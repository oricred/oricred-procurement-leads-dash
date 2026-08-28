from io import BytesIO

import pytest
from openpyxl import Workbook

from app.schemas.contact import ContactCreate
from app.services.lead_contact_import import (
    ImportRow,
    _canonicalise_row,
    _metadata_notes,
    parse_import_file,
)


def test_parse_csv_enriched_lead() -> None:
    rows = parse_import_file(
        "enriched.csv",
        b"lead_id,company,contact_email,contact_phone\n"
        b"lead-1,Example Ltd,person@example.com,012 345 6789\n",
    )
    assert len(rows) == 1
    assert rows[0].row_number == 2
    assert rows[0].values["lead_id"] == "lead-1"
    assert rows[0].values["contact_email"] == "person@example.com"


def test_parse_xlsx_enriched_lead() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["lead_id", "company", "contact_phone"])
    sheet.append(["lead-2", "Example Ltd", "021 555 0000"])
    content = BytesIO()
    workbook.save(content)
    rows = parse_import_file("enriched.xlsx", content.getvalue())
    assert rows[0].values["lead_id"] == "lead-2"
    assert rows[0].values["contact_phone"] == "021 555 0000"


def test_import_metadata_is_saved_in_notes() -> None:
    row = parse_import_file(
        "enriched.csv",
        b"lead_id,contact_source_url,contact_confidence,enrichment_date,research_notes\n"
        b"lead-1,https://example.com,High,2026-07-18,Public contact\n",
    )[0]
    notes = _metadata_notes(row)
    assert notes and "Source URL: https://example.com" in notes
    assert "Research notes: Public contact" in notes


def test_phone_only_contact_schema_is_supported() -> None:
    contact = ContactCreate(first_name="Example", last_name="Ltd", phone_direct="021 555 0000")
    assert contact.email is None


def test_canonicalise_rows_maps_email_alias() -> None:
    row = ImportRow(2, {"lead_id": "l1", "company": "Acme", "email": "a@b.com", "phone": "123"})
    canonical = _canonicalise_row(row)
    assert canonical.values["contact_email"] == "a@b.com"
    assert canonical.values["contact_phone"] == "123"
    assert canonical.values["lead_id"] == "l1"
    assert canonical.values["company"] == "Acme"


def test_canonicalise_rows_does_not_overwrite_existing() -> None:
    row = ImportRow(2, {"contact_email": "existing@b.com", "email": "alias@b.com"})
    canonical = _canonicalise_row(row)
    assert canonical.values["contact_email"] == "existing@b.com"


def test_canonicalise_rows_phone_aliases() -> None:
    for alias in ("telephone", "tel", "cell", "mobile"):
        row = ImportRow(2, {"lead_id": "l1", alias: "021 555 0000"})
        canonical = _canonicalise_row(row)
        assert canonical.values.get("contact_phone") == "021 555 0000", f"alias {alias} not mapped"


def test_canonicalise_rows_name_aliases() -> None:
    for alias in ("name", "contact_person", "full_name"):
        row = ImportRow(2, {"lead_id": "l1", alias: "Jane Doe"})
        canonical = _canonicalise_row(row)
        assert canonical.values.get("contact_name") == "Jane Doe", f"alias {alias} not mapped"


def test_canonicalise_rows_job_title_aliases() -> None:
    for alias in ("job_title", "title", "position"):
        row = ImportRow(2, {"lead_id": "l1", alias: "Manager"})
        canonical = _canonicalise_row(row)
        assert canonical.values.get("contact_job_title") == "Manager", f"alias {alias} not mapped"


def test_canonicalise_rows_company_aliases() -> None:
    for alias in ("company_name", "supplier", "supplier_name"):
        row = ImportRow(2, {"lead_id": "l1", alias: "Acme Corp"})
        canonical = _canonicalise_row(row)
        assert canonical.values.get("company") == "Acme Corp", f"alias {alias} not mapped"


@pytest.mark.parametrize("header_alias", ["email", "e-mail", "mail"])
def test_parse_csv_with_email_alias(header_alias: str) -> None:
    csv_content = f"company,{header_alias}\nAcme Ltd,jane@acme.com\n".encode()
    rows = parse_import_file("contacts.csv", csv_content)
    assert len(rows) == 1
    assert rows[0].values["company"] == "Acme Ltd"
    assert rows[0].values["contact_email"] == "jane@acme.com"


@pytest.mark.parametrize("header_alias", ["phone", "telephone", "tel", "cell", "mobile"])
def test_parse_csv_with_phone_alias(header_alias: str) -> None:
    csv_content = f"lead_id,{header_alias}\nl1,072 123 4567\n".encode()
    rows = parse_import_file("contacts.csv", csv_content)
    assert rows[0].values["contact_phone"] == "072 123 4567"


def test_parse_xlsx_with_email_only() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["company", "email"])
    sheet.append(["Acme Ltd", "jane@acme.com"])
    content = BytesIO()
    workbook.save(content)
    rows = parse_import_file("enriched.xlsx", content.getvalue())
    assert rows[0].values["company"] == "Acme Ltd"
    assert rows[0].values["contact_email"] == "jane@acme.com"


def test_parse_xlsx_with_company_name_alias() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["supplier", "contact_email"])
    sheet.append(["Acme Ltd", "jane@acme.com"])
    content = BytesIO()
    workbook.save(content)
    rows = parse_import_file("enriched.xlsx", content.getvalue())
    assert rows[0].values["company"] == "Acme Ltd"
    assert rows[0].values["contact_email"] == "jane@acme.com"


class TestAmbiguousCompanyNames:
    """Regression guard for the H4 defect.

    _find_company_and_opportunity called scalar_one_or_none() on an ilike match.
    Two companies sharing a name raised MultipleResultsFound, which nothing
    handled — an opaque HTTP 500 partway through an operator's import. Duplicate
    supplier names are routine: the award pipeline creates `provisional:`
    companies and the historical sync creates `historical:` ones alongside
    canonical rows with the same name.
    """

    @staticmethod
    async def _companies(session_factory, rows):
        from app.models.company import Company

        async with session_factory() as db:
            for api_id, name in rows:
                db.add(Company(api_id=api_id, name=name))
            await db.commit()

    async def test_two_real_companies_skip_the_row_instead_of_raising(self, import_db):
        from app.services.lead_contact_import import _find_company_and_opportunity

        await self._companies(import_db, [
            ("tsa-1", "ABC Trading"),
            ("tsa-2", "ABC Trading"),
        ])
        async with import_db() as db:
            company, _opp, error = await _find_company_and_opportunity("ABC Trading", db)

        assert company is None
        assert error is not None and "2 companies match" in error

    async def test_a_canonical_row_wins_over_a_provisional_stub(self, import_db):
        from app.services.lead_contact_import import _find_company_and_opportunity

        await self._companies(import_db, [
            ("tsa-1", "ABC Trading"),
            ("provisional:award-9", "ABC Trading"),
        ])
        async with import_db() as db:
            company, _opp, error = await _find_company_and_opportunity("ABC Trading", db)

        assert error is None
        assert company is not None and company.api_id == "tsa-1"

    async def test_a_unique_name_resolves(self, import_db):
        from app.services.lead_contact_import import _find_company_and_opportunity

        await self._companies(import_db, [("tsa-1", "Sizwe Construction")])
        async with import_db() as db:
            company, _opp, error = await _find_company_and_opportunity("Sizwe Construction", db)

        assert error is None
        assert company is not None and company.api_id == "tsa-1"

    async def test_wildcards_in_a_name_are_matched_literally(self, import_db):
        """ilike treated % and _ as wildcards; a supplier named 'A_B' matched 'AxB'."""
        from app.services.lead_contact_import import _find_company_and_opportunity

        await self._companies(import_db, [("tsa-1", "AxB Trading")])
        async with import_db() as db:
            company, _opp, error = await _find_company_and_opportunity("A_B Trading", db)

        assert company is None
        assert error is None  # simply unknown, not ambiguous

    async def test_an_unknown_name_is_not_an_error(self, import_db):
        from app.services.lead_contact_import import _find_company_and_opportunity

        async with import_db() as db:
            company, _opp, error = await _find_company_and_opportunity("Nobody Ltd", db)
        assert company is None and error is None


class TestImportRowLimit:
    def test_a_file_over_the_row_cap_is_rejected(self):
        import io as _io

        from app.services.lead_contact_import import MAX_IMPORT_ROWS, parse_import_file

        buf = _io.StringIO()
        buf.write("company,contact_email\n")
        for i in range(MAX_IMPORT_ROWS + 1):
            buf.write(f"Company {i},p{i}@example.com\n")

        with pytest.raises(ValueError, match="Split it into files"):
            parse_import_file("big.csv", buf.getvalue().encode("utf-8"))

    def test_a_file_at_the_cap_is_accepted(self):
        import io as _io

        from app.services.lead_contact_import import MAX_IMPORT_ROWS, parse_import_file

        buf = _io.StringIO()
        buf.write("company,contact_email\n")
        for i in range(MAX_IMPORT_ROWS):
            buf.write(f"Company {i},p{i}@example.com\n")

        assert len(parse_import_file("ok.csv", buf.getvalue().encode("utf-8"))) == MAX_IMPORT_ROWS
