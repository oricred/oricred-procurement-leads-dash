import csv
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.company import Company
from app.models.contact import Contact
from app.models.opportunity import Opportunity
from app.services.lead_scoring import refresh_lead_scoring
from app.services.text_utils import normalise_company_name

IMPORT_SOURCE = "lead_import"
# Each row costs two or three queries to decide. Bounding the row count keeps a
# single request from monopolising a worker; the file-size ceiling lives in the
# route (see _read_bounded in app/api/leads.py).
MAX_IMPORT_ROWS = 10_000
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
IMPORT_METADATA_FIELDS = (
    ("contact_source_url", "Source URL"),
    ("contact_confidence", "Confidence"),
    ("enrichment_date", "Enrichment date"),
    ("research_notes", "Research notes"),
)

COLUMN_ALIASES = {
    "email": "contact_email",
    "e-mail": "contact_email",
    "mail": "contact_email",
    "phone": "contact_phone",
    "telephone": "contact_phone",
    "tel": "contact_phone",
    "cell": "contact_phone",
    "cellphone": "contact_phone",
    "mobile": "contact_phone",
    "phone_direct": "contact_phone",
    "phone_mobile": "contact_phone",
    "name": "contact_name",
    "contact_person": "contact_name",
    "full_name": "contact_name",
    "person": "contact_name",
    "job_title": "contact_job_title",
    "title": "contact_job_title",
    "position": "contact_job_title",
    "designation": "contact_job_title",
    "company_name": "company",
    "supplier": "company",
    "supplier_name": "company",
}


@dataclass
class ImportRow:
    row_number: int
    values: dict[str, str]


@dataclass
class ImportDecision:
    row: ImportRow
    action: str
    message: str
    opportunity: Opportunity | None = None
    contact: Contact | None = None
    company_id: str | None = None


def _clean(value: object | None) -> str:
    return str(value).strip() if value is not None else ""


def _normalise(value: str) -> str:
    return " ".join(value.lower().split())


def _normalise_phone(value: str) -> str:
    return re.sub(r"\D", "", value)


def _split_name(name: str) -> tuple[str, str]:
    parts = name.split(None, 1)
    return (parts[0], parts[1] if len(parts) > 1 else "")


def _canonicalise_row(row: ImportRow) -> ImportRow:
    mapped = row.values.copy()
    for alias, canonical in COLUMN_ALIASES.items():
        if alias in mapped and canonical not in mapped:
            mapped[canonical] = mapped[alias]
    return ImportRow(row_number=row.row_number, values=mapped)


SYNTHETIC_COMPANY_PREFIXES = ("provisional:", "historical:", "award:")


async def _find_company_and_opportunity(
    company_name: str, db: AsyncSession
) -> tuple[Company | None, Opportunity | None, str | None]:
    """Resolve a company by name.

    Returns (company, opportunity, error). `error` is set when the name matches
    more than one real company, so the caller can skip the row with a useful
    message. The previous implementation called scalar_one_or_none() on an
    ilike match and raised MultipleResultsFound — an unhandled HTTP 500 partway
    through an import. Duplicate supplier names are routine here: the award
    pipeline creates `provisional:` companies and the historical sync creates
    `historical:` ones, both alongside canonical rows with the same name.
    """
    target = normalise_company_name(company_name)
    if not target:
        return None, None, "Blank company name"

    # Exact match first — indexed, and unambiguous when it hits.
    candidates = list(
        (
            await db.execute(
                select(Company).where(func.lower(Company.name) == company_name.strip().lower())
            )
        )
        .scalars()
        .all()
    )
    if not candidates:
        # Fall back to the same normalisation the enrichment matcher uses, so
        # the two agree about what "the same company" means.
        candidates = [
            c
            for c in (await db.execute(select(Company))).scalars().all()
            if normalise_company_name(c.name) == target
        ]

    if not candidates:
        return None, None, None
    if len(candidates) > 1:
        real = [
            c for c in candidates
            if not (c.api_id or "").startswith(SYNTHETIC_COMPANY_PREFIXES)
        ]
        if len(real) != 1:
            return None, None, (
                f"{len(candidates)} companies match {company_name!r} — "
                "import by lead_id instead"
            )
        candidates = real

    company = candidates[0]
    opportunity = (
        (
            await db.execute(
                select(Opportunity)
                .where(Opportunity.company_id == company.id)
                .order_by(Opportunity.created_at.desc())
                .limit(1)
            )
        )
        .scalars()
        .first()
    )
    return company, opportunity, None


def parse_import_file(filename: str | None, content: bytes) -> list[ImportRow]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        try:
            reader = csv.DictReader(StringIO(content.decode("utf-8-sig")))
            rows = [
                ImportRow(index, {_clean(key): _clean(value) for key, value in raw.items() if key})
                for index, raw in enumerate(reader, start=2)
            ]
        except UnicodeDecodeError as exc:
            raise ValueError("CSV files must be UTF-8 encoded") from exc
        except csv.Error as exc:
            # A malformed CSV is the operator's problem to fix, not a 500.
            raise ValueError(f"The CSV file could not be read: {exc}") from exc
    elif suffix == ".xlsx":
        try:
            worksheet = load_workbook(BytesIO(content), read_only=True, data_only=True).active
            rows_iter = worksheet.iter_rows(values_only=True)
            headers = [_clean(value) for value in next(rows_iter, ())]
            rows = [
                ImportRow(
                    index,
                    {
                        headers[column]: _clean(value)
                        for column, value in enumerate(raw)
                        if column < len(headers) and headers[column]
                    },
                )
                for index, raw in enumerate(rows_iter, start=2)
            ]
        except Exception as exc:
            raise ValueError("The XLSX file could not be read") from exc
    else:
        raise ValueError("Only .csv and .xlsx files are supported")

    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(
            f"That file has {len(rows):,} rows. Split it into files of "
            f"{MAX_IMPORT_ROWS:,} rows or fewer."
        )
    return [_canonicalise_row(row) for row in rows]


def _metadata_notes(row: ImportRow) -> str | None:
    details = [
        f"{label}: {row.values[field]}"
        for field, label in IMPORT_METADATA_FIELDS
        if row.values.get(field)
    ]
    return "[Lead import enrichment]\n" + "\n".join(details) if details else None


async def _decide(row: ImportRow, db: AsyncSession) -> ImportDecision:
    lead_id = row.values.get("lead_id", "")
    company_name = row.values.get("company", "")

    opportunity = None
    company_id = None

    if lead_id:
        opportunity = await db.get(Opportunity, lead_id)
        if not opportunity:
            return ImportDecision(row, "skip", "Unknown lead_id")
        if not opportunity.company_id:
            return ImportDecision(row, "skip", "Lead has no linked company", opportunity)
        company_id = opportunity.company_id
    elif company_name:
        company, opportunity, error = await _find_company_and_opportunity(company_name, db)
        if error:
            return ImportDecision(row, "skip", error)
        if not company:
            return ImportDecision(row, "skip", f"Unknown company: {company_name}")
        company_id = company.id
        if not company_id:
            return ImportDecision(row, "skip", "Company has no id")
    else:
        return ImportDecision(row, "skip", "Missing lead_id or company")

    email = row.values.get("contact_email", "").lower()
    phone = row.values.get("contact_phone", "")
    if not email and not phone:
        return ImportDecision(
            row, "skip", "No email or phone to import", opportunity, None, company_id
        )
    if email and not EMAIL_PATTERN.match(email):
        return ImportDecision(row, "skip", "Invalid contact_email", opportunity, None, company_id)

    contacts = (
        (await db.execute(select(Contact).where(Contact.company_id == company_id)))
        .scalars()
        .all()
    )
    imported_contacts = [contact for contact in contacts if contact.source == IMPORT_SOURCE]
    if email:
        matching = next(
            (
                contact
                for contact in imported_contacts
                if _normalise(contact.email or "") == _normalise(email)
            ),
            None,
        )
        protected = next(
            (
                contact
                for contact in contacts
                if contact.source != IMPORT_SOURCE
                and _normalise(contact.email or "") == _normalise(email)
            ),
            None,
        )
        if protected:
            return ImportDecision(
                row, "skip", "Email belongs to a protected existing contact", opportunity, None, company_id
            )
        if matching:
            return ImportDecision(row, "update", "Update imported contact", opportunity, matching, company_id)
    if phone:
        matching = next(
            (
                contact
                for contact in imported_contacts
                if _normalise_phone(contact.phone_direct or contact.phone_mobile or "")
                == _normalise_phone(phone)
            ),
            None,
        )
        if matching:
            return ImportDecision(row, "update", "Update imported contact", opportunity, matching, company_id)
    name = row.values.get("contact_name", "")
    if name:
        first_name, last_name = _split_name(name)
        matching = next(
            (
                contact
                for contact in imported_contacts
                if _normalise(contact.first_name) == _normalise(first_name)
                and _normalise(contact.last_name) == _normalise(last_name)
            ),
            None,
        )
        if matching:
            return ImportDecision(row, "update", "Update imported contact", opportunity, matching, company_id)
    return ImportDecision(row, "create", "Create imported contact", opportunity, None, company_id)


def _result(decision: ImportDecision) -> dict[str, object]:
    return {
        "row": decision.row.row_number,
        "lead_id": decision.row.values.get("lead_id") or None,
        "company": decision.row.values.get("company") or None,
        "action": decision.action,
        "message": decision.message,
    }


def _validate_columns(rows: list[ImportRow]) -> None:
    if not rows:
        raise ValueError("The import file contains no data rows")
    has_lead_id = any("lead_id" in row.values for row in rows)
    has_company = any("company" in row.values for row in rows)
    if not has_lead_id and not has_company:
        raise ValueError("The import file must include a lead_id or company column")


def _row_identity(decision: ImportDecision) -> tuple[str, str] | None:
    """A key for detecting the same contact appearing twice in one file."""
    if not decision.company_id:
        return None
    values = decision.row.values
    email = values.get("contact_email", "").strip().lower()
    phone = _normalise_phone(values.get("contact_phone", ""))
    name = _normalise(values.get("contact_name", ""))
    return (decision.company_id, email or phone or name)


async def _decide_all(rows: list[ImportRow], db: AsyncSession) -> list[ImportDecision]:
    """Decide every row once, against committed state.

    apply_import used to call preview_import and then re-decide each row, which
    doubled the query cost and let the two passes disagree: autoflush made rows
    written during the loop visible to the second pass, so the counts shown to
    the operator did not match what was applied.

    no_autoflush keeps the decisions consistent with the state the operator saw
    in the preview.
    """
    _validate_columns(rows)
    decisions: list[ImportDecision] = []
    seen: set[tuple[str, str]] = set()
    with db.no_autoflush:
        for row in rows:
            decision = await _decide(row, db)
            identity = _row_identity(decision)
            if decision.action != "skip" and identity is not None:
                if identity in seen:
                    decision = ImportDecision(
                        row, "skip", "Duplicate of an earlier row in this file"
                    )
                else:
                    seen.add(identity)
            decisions.append(decision)
    return decisions


def _summarise(decisions: list[ImportDecision]) -> dict[str, object]:
    return {
        "total_rows": len(decisions),
        "creates": sum(decision.action == "create" for decision in decisions),
        "updates": sum(decision.action == "update" for decision in decisions),
        "skips": sum(decision.action == "skip" for decision in decisions),
        "rows": [_result(decision) for decision in decisions],
    }


async def preview_import(rows: list[ImportRow], db: AsyncSession) -> dict[str, object]:
    return _summarise(await _decide_all(rows, db))


async def apply_import(rows: list[ImportRow], db: AsyncSession) -> dict[str, object]:
    decisions = await _decide_all(rows, db)
    preview = _summarise(decisions)
    affected_company_ids: set[str] = set()
    applied = 0
    for decision in decisions:
        row = decision.row
        if decision.action == "skip":
            continue
        company_id = decision.company_id or (
            decision.opportunity.company_id if decision.opportunity else None
        )
        if not company_id:
            continue
        values = row.values
        contact_name = values.get("contact_name", "")
        job_title = values.get("contact_job_title", "")
        email = values.get("contact_email", "").lower()
        phone = values.get("contact_phone", "")
        notes = _metadata_notes(row)
        if decision.action == "create":
            first_name, last_name = _split_name(
                contact_name or values.get("company", "") or "Unknown contact"
            )
            has_primary = bool(
                (
                    await db.execute(
                        select(Contact.id)
                        .where(
                            Contact.company_id == company_id,
                            Contact.is_primary.is_(True),
                        )
                        .limit(1)
                    )
                ).scalar_one_or_none()
            )
            decision.contact = Contact(
                company_id=company_id,
                first_name=first_name,
                last_name=last_name,
                job_title=job_title or None,
                email=email or None,
                phone_direct=phone or None,
                is_primary=not has_primary,
                notes=notes,
                source=IMPORT_SOURCE,
            )
            db.add(decision.contact)
        elif decision.contact:
            if contact_name:
                decision.contact.first_name, decision.contact.last_name = _split_name(contact_name)
            if job_title:
                decision.contact.job_title = job_title
            if email:
                decision.contact.email = email
            if phone:
                decision.contact.phone_direct = phone
            if notes:
                decision.contact.notes = notes
        affected_company_ids.add(company_id)
        applied += 1

    await db.flush()
    for company_id in affected_company_ids:
        opportunities = (
            (await db.execute(select(Opportunity).where(Opportunity.company_id == company_id)))
            .scalars()
            .all()
        )
        for opportunity in opportunities:
            await refresh_lead_scoring(opportunity, db)
    await db.commit()
    return {**preview, "applied": applied}
